import os
import argparse
import numpy as np
import tensorflow as tf
import gc
import logging
import pandas as pd
import time
import pickle
from strategy.FedAvg import FedAvg
from strategy.FedAvg2 import FedAvg2
from strategy.FedAGRU import FedAGRU
from strategy.FedProx import FedProx
from strategy.FedProxAdaptive import FedProxAdaptive
from strategy.SCAFFOLD import SCAFFOLD
from models import dense, dense_sim, gru, fedprox_wrapper, fedprox_adaptive_wrapper, scaffold_wrapper
from sklearn.metrics import f1_score, precision_score, recall_score, confusion_matrix, classification_report
# Control GPU memory usage
gpus = tf.config.experimental.list_physical_devices('GPU')
if gpus:
    try:
        tf.config.experimental.set_memory_growth(gpus[0], True)
        gpu_options = tf.compat.v1.GPUOptions(
            per_process_gpu_memory_fraction=0.9,
            allow_growth=True,
            polling_active_delay=10,
            allocator_type='BFC'
        )
        config = tf.compat.v1.ConfigProto(gpu_options=gpu_options)
        config.gpu_options.allow_growth = True
        tf.compat.v1.keras.backend.set_session(tf.compat.v1.Session(config=config))
    except Exception as e:
        print(f"GPU configuration error: {e}")

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    PURPLE = '\033[95m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'

def setup_logger(n_clients, partition_type, strategy="FedUp"):
    log_filename = f"{strategy}_{n_clients}client_{partition_type}.log"
    logging.basicConfig(
        filename=log_filename,
        level=logging.INFO,
        format='%(levelname)s: %(message)s',
        filemode='w'
    )
    print(bcolors.OKCYAN + f"Logging to {log_filename}" + bcolors.ENDC)
    return logging.getLogger(), log_filename

def parse_partition_type(partition_type):
    if '-' in partition_type:
        parts = partition_type.split('-')
        if len(parts) == 2:
            return parts[0], int(parts[1])
    raise ValueError(f"Invalid partition type '{partition_type}'. Use format 'iid-10'")

def setup_paths(client_id, partition_type, client_count):
    partitions_root = os.path.join("data", "partitions")
    client_folder = f"{client_count}_client"
    partition_dir = os.path.join(partitions_root, client_folder, partition_type)
    
    return {
        'train_X': os.path.join(partition_dir, f"client_{client_id}_X_train.npy"),
        'train_y': os.path.join(partition_dir, f"client_{client_id}_y_train.npy"),
        'test_X': os.path.join("data", "X_test.npy"),
        'test_y': os.path.join("data", "y_test.npy")
    }

def create_model(input_dim, num_classes, model_type, batch_size=8192, strategy=None, client_id=None):
    if strategy and hasattr(strategy, 'name'):
        if strategy.name == "FedProx":
            if model_type.lower() == "dense":
                return fedprox_wrapper.create_fedprox_dense_model(input_dim, num_classes, batch_size, strategy)
            elif model_type.lower() in ["gru"]:
                return fedprox_wrapper.create_fedprox_gru_model((1, input_dim), num_classes, batch_size, strategy)
        elif strategy.name == "FedProxAdaptive":
            if model_type.lower() == "dense":
                return fedprox_adaptive_wrapper.create_fedprox_adaptive_dense_model(input_dim, num_classes, batch_size, strategy, client_id)
            elif model_type.lower() in ["gru"]:
                return fedprox_adaptive_wrapper.create_fedprox_adaptive_gru_model((1, input_dim), num_classes, batch_size, strategy, client_id)
        elif strategy.name == "SCAFFOLD":
            if model_type.lower() == "dense":
                return scaffold_wrapper.create_scaffold_dense_model(input_dim, num_classes, batch_size, strategy, client_id)
            elif model_type.lower() in ["gru"]:
                return scaffold_wrapper.create_scaffold_gru_model((1, input_dim), num_classes, batch_size, strategy, client_id)
    
    if model_type.lower() == "dense":
        return dense.create_enhanced_dense_model(input_dim, num_classes, batch_size)
    elif model_type.lower() in ["gru"]:
        return gru.create_enhanced_gru_model((1, input_dim), num_classes, batch_size)
    else:
        raise ValueError(f"Unknown model type: {model_type}")

def create_client_dataset(X_path, y_path, input_dim, num_classes, batch_size, is_gru=False):
    def generator():
        X_mmap = np.load(X_path, mmap_mode='r')
        y_mmap = np.load(y_path, mmap_mode='r')
        total_samples = X_mmap.shape[0]
        chunk_size = 10000
        
        for start_idx in range(0, total_samples, chunk_size):
            end_idx = min(start_idx + chunk_size, total_samples)
            X_chunk = np.array(X_mmap[start_idx:end_idx])
            
            if is_gru:
                X_chunk = X_chunk.reshape(-1, 1, input_dim)
            
            y_chunk = np.array(y_mmap[start_idx:end_idx])
            if num_classes and (len(y_chunk.shape) == 1 or y_chunk.shape[1] == 1):
                y_chunk = tf.keras.utils.to_categorical(y_chunk, num_classes=num_classes)
            
            yield X_chunk, y_chunk
            del X_chunk, y_chunk
    
    if is_gru:
        output_signature = (
            tf.TensorSpec(shape=(None, 1, input_dim), dtype=tf.float32),
            tf.TensorSpec(shape=(None, num_classes), dtype=tf.float32)
        )
    else:
        output_signature = (
            tf.TensorSpec(shape=(None, input_dim), dtype=tf.float32),
            tf.TensorSpec(shape=(None, num_classes), dtype=tf.float32)
        )
    
    dataset = tf.data.Dataset.from_generator(generator, output_signature=output_signature)
    return dataset.unbatch().batch(batch_size).prefetch(tf.data.AUTOTUNE)

def aggressive_memory_cleanup():
    tf.keras.backend.clear_session()
    gc.collect()
    for _ in range(3):
        gc.collect()
    time.sleep(1)

def evaluate_model_with_metrics(model, test_dataset, num_classes, class_names=None, round_num=None, strategy_name=None, partition_type=None):
    # Get predictions and true labels
    all_predictions = model.predict(test_dataset, verbose=1)
    y_pred = np.argmax(all_predictions, axis=1)
    
    y_true = []
    for batch_x, batch_y in test_dataset:
        if len(batch_y.shape) > 1 and batch_y.shape[1] > 1:
            batch_y_true = np.argmax(batch_y.numpy(), axis=1)
        else:
            batch_y_true = batch_y.numpy().astype(int)
        y_true.extend(batch_y_true)
    
    y_true = np.array(y_true)
    
    # Ensure sizes match
    min_size = min(len(y_true), len(y_pred))
    y_true = y_true[:min_size]
    y_pred = y_pred[:min_size]
    
    # Get basic metrics
    test_results = model.evaluate(test_dataset, verbose=0)
    test_loss, accuracy = test_results[0], test_results[1]
    
    # Calculate macro averages (better for multiclass)
    f1_macro = f1_score(y_true, y_pred, average='macro', zero_division=0)
    precision_macro = precision_score(y_true, y_pred, average='macro', zero_division=0)
    recall_macro = recall_score(y_true, y_pred, average='macro', zero_division=0)
    
    # Per-class metrics and confusion matrix
    f1_per_class = f1_score(y_true, y_pred, average=None, zero_division=0)
    precision_per_class = precision_score(y_true, y_pred, average=None, zero_division=0)
    recall_per_class = recall_score(y_true, y_pred, average=None, zero_division=0)
    cm = confusion_matrix(y_true, y_pred)
    
    if class_names is None:
        class_names = [f"Class_{i}" for i in range(num_classes)]
    
    class_report = classification_report(y_true, y_pred, target_names=class_names, zero_division=0)
    
    # Log detailed metrics to files
    if round_num is not None and strategy_name is not None:
        log_detailed_metrics_to_files(round_num, strategy_name, partition_type, cm, class_report, 
                                    class_names, f1_per_class, precision_per_class, recall_per_class,
                                    f1_macro, precision_macro, recall_macro)
    
    return (test_loss, accuracy, f1_macro, precision_macro, recall_macro, 
            (f1_per_class, precision_per_class, recall_per_class), cm)

def log_detailed_metrics_to_files(round_num, strategy_name, partition_type, confusion_matrix, class_report, 
                                 class_names, f1_per_class, precision_per_class, recall_per_class,
                                 f1_macro, precision_macro, recall_macro):
    results_dir = "results"
    os.makedirs(results_dir, exist_ok=True)
    
    detailed_log_filename = os.path.join(results_dir, f"{strategy_name}_{partition_type}_detailed_metrics.log")
    
    with open(detailed_log_filename, 'a') as f:
        f.write(f"\n{'='*60}\n")
        f.write(f"ROUND {round_num} - DETAILED METRICS\n")
        f.write(f"Strategy: {strategy_name} | Partition: {partition_type}\n")
        f.write(f"{'='*60}\n")
        
        f.write(f"\nMACRO AVERAGES:\n")
        f.write(f"F1={f1_macro:.4f}, Precision={precision_macro:.4f}, Recall={recall_macro:.4f}\n")
        
        f.write(f"\nPER-CLASS METRICS:\n")
        f.write(f"{'Class Name':<15} {'F1':<8} {'Precision':<10} {'Recall':<8}\n")
        f.write(f"{'-'*50}\n")
        for i in range(len(f1_per_class)):
            class_name = class_names[i] if i < len(class_names) else f"Class_{i}"
            f.write(f"{class_name[:15]:<15} {f1_per_class[i]:<8.3f} {precision_per_class[i]:<10.3f} {recall_per_class[i]:<8.3f}\n")
        
        # SIMPLIFIED: Just reference the separate confusion matrix file
        f.write(f"\nCONFUSION MATRIX:\n")
        f.write(f"Full confusion matrix saved to separate CSV file:\n")
        f.write(f"{strategy_name}_{partition_type}_round_{round_num}_confusion_matrix_full.csv\n")
        f.write(f"Matrix shape: {confusion_matrix.shape}\n")
        
        f.write(f"\nCLASSIFICATION REPORT:\n")
        f.write(class_report)
        f.write(f"\n{'='*60}\n")
    
    save_detailed_metrics(round_num, strategy_name, partition_type, confusion_matrix, class_report, 
                         class_names, f1_per_class, precision_per_class, recall_per_class)

def save_confusion_matrix_separately(round_num, strategy_name, partition_type, confusion_matrix, class_names):
    """Save confusion matrix to a separate CSV file with full class names."""
    results_dir = "results"
    os.makedirs(results_dir, exist_ok=True)
    
    # Create filename for confusion matrix
    cm_filename = os.path.join(results_dir, f"{strategy_name}_{partition_type}_round_{round_num}_confusion_matrix_full.csv")
    
    # Ensure we have the right number of class names
    num_classes = confusion_matrix.shape[0]
    if len(class_names) < num_classes:
        # Extend class_names if needed
        extended_class_names = class_names.copy()
        for i in range(len(class_names), num_classes):
            extended_class_names.append(f"Class_{i}")
        class_names = extended_class_names
    elif len(class_names) > num_classes:
        # Truncate class_names if too many
        class_names = class_names[:num_classes]
    
    # Create DataFrame with full class names (no truncation)
    cm_df = pd.DataFrame(confusion_matrix, 
                        index=[f"{i}_{name}" for i, name in enumerate(class_names)], 
                        columns=[f"{i}_{name}" for i, name in enumerate(class_names)])
    
    # Save to CSV
    cm_df.to_csv(cm_filename)
    
    # Also save a summary text file
    summary_filename = os.path.join(results_dir, f"{strategy_name}_{partition_type}_round_{round_num}_confusion_summary.txt")
    with open(summary_filename, 'w') as f:
        f.write(f"Confusion Matrix Summary - Round {round_num}\n")
        f.write(f"Strategy: {strategy_name} | Partition: {partition_type}\n")
        f.write(f"{'='*60}\n\n")
        
        f.write(f"Matrix Shape: {confusion_matrix.shape}\n")
        f.write(f"Total Classes: {num_classes}\n\n")
        
        f.write("Class Names:\n")
        for i, name in enumerate(class_names):
            f.write(f"  {i}: {name}\n")
        
        f.write(f"\nConfusion Matrix saved to: {cm_filename}\n")
    
    print(f"Confusion matrix saved to: {cm_filename}")
    print(f"Summary saved to: {summary_filename}")

def save_detailed_metrics(round_num, strategy_name, partition_type, confusion_matrix, class_report, 
                         class_names, f1_per_class, precision_per_class, recall_per_class):
    results_dir = "results"
    os.makedirs(results_dir, exist_ok=True)
    
    filename_prefix = f"{strategy_name}_{partition_type}_round_{round_num}"
    
    # Save confusion matrix separately (with full class names)
    save_confusion_matrix_separately(round_num, strategy_name, partition_type, confusion_matrix, class_names)
    
    # Save per-class metrics
    per_class_filename = os.path.join(results_dir, f"{filename_prefix}_per_class_metrics.csv")
    
    # Ensure we have the right number of class names for per-class metrics
    num_metrics = len(f1_per_class)
    if len(class_names) < num_metrics:
        extended_class_names = class_names.copy()
        for i in range(len(class_names), num_metrics):
            extended_class_names.append(f"Class_{i}")
        class_names_for_metrics = extended_class_names
    else:
        class_names_for_metrics = class_names[:num_metrics]
    
    per_class_df = pd.DataFrame({
        'Class_Index': range(num_metrics),
        'Class_Name': class_names_for_metrics,
        'F1_Score': f1_per_class,
        'Precision': precision_per_class,
        'Recall': recall_per_class
    })
    per_class_df.to_csv(per_class_filename, index=False)

def create_enhanced_excel_report(excel_filename, main_results_df, per_class_metrics, 
                               class_names, current_round, confusion_matrix):
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill
    
    try:
        wb = openpyxl.load_workbook(excel_filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    
    # Main results sheet
    if 'Overall_Metrics' in wb.sheetnames:
        wb.remove(wb['Overall_Metrics'])
    ws_main = wb.create_sheet('Overall_Metrics')
    
    for r in dataframe_to_rows(main_results_df, index=False, header=True):
        ws_main.append(r)
    
    for cell in ws_main[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # REMOVED: Confusion matrix section - no longer added to Excel
    
    # Per-class metrics sheet
    if per_class_metrics is not None:
        sheet_name = f'Round_{current_round}_PerClass'
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws_per_class = wb.create_sheet(sheet_name)
        
        f1_per_class, precision_per_class, recall_per_class = per_class_metrics
        per_class_df = pd.DataFrame({
            'Class': class_names,
            'F1_Score': f1_per_class,
            'Precision': precision_per_class,
            'Recall': recall_per_class
        })
        
        for r in dataframe_to_rows(per_class_df, index=False, header=True):
            ws_per_class.append(r)
    
    wb.save(excel_filename)

def main():
    parser = argparse.ArgumentParser(description="Federated Learning Simulation")
    parser.add_argument("--n_clients", type=int, default=5, help="Number of clients to simulate")
    parser.add_argument("--partition_type", type=str, default="iid-10", help="Partition type (e.g., 'iid-10')")
    parser.add_argument("--model_type", type=str, default="gru", help="Model type: 'dense' or 'gru'")
    parser.add_argument("--rounds", type=int, default=10, help="Number of federated learning rounds")
    parser.add_argument("--strategy", type=str, default="FedAvg", help="Strategy: 'FedAvg', 'FedAvg2', 'FedAGRU', 'FedProx', 'FedProxAdaptive', or 'SCAFFOLD'")
    parser.add_argument("--batch_size", type=int, default=8192, help="Batch size for training")
    parser.add_argument("--epochs", type=int, default=5, help="Number of epochs per round")
    parser.add_argument("--weights_cache_dir", type=str, default="temp_weights", help="Directory to cache weights")
    parser.add_argument("--patience_T", type=int, default=3, help="Patience rounds for FedAGRU")
    parser.add_argument("--disable_client_selection", action='store_true', help="Disable client selection based on importance")
    parser.add_argument("--mu", type=float, default=0.01, help="Proximal term coefficient for FedProx")
    parser.add_argument("--adaptive_mu", action='store_true', help="Use adaptive mu for FedProx")
    parser.add_argument("--warmup_rounds", type=int, default=3, help="Warmup rounds for adaptive mu")
    parser.add_argument("--delayed_start", action='store_true', help="Delay FedProx for first 2 rounds")
    
    args = parser.parse_args()
    
    os.makedirs(args.weights_cache_dir, exist_ok=True)
    
    if args.strategy == "FedAGRU":
        threshold_v = 1.0 / args.n_clients
        strategy = FedAGRU()
        print(f"{bcolors.PURPLE}Using FedAGRU with threshold_v={threshold_v:.4f}, patience_T={args.patience_T}{bcolors.ENDC}")
    elif args.strategy == "FedAvg2":
        strategy = FedAvg2()
        print(f"{bcolors.PURPLE}Using FedAvg2 with loss-weighted aggregation{bcolors.ENDC}")
    elif args.strategy == "FedProx":
        strategy = FedProx(mu=args.mu, adaptive_mu=args.adaptive_mu, warmup_rounds=args.warmup_rounds, delayed_start=args.delayed_start)
        print(f"{bcolors.PURPLE}Using FedProx with base_mu={args.mu}, adaptive={args.adaptive_mu}, delayed_start={args.delayed_start}{bcolors.ENDC}")
    elif args.strategy == "FedProxAdaptive":
        strategy = FedProxAdaptive(base_mu=args.mu)
        print(f"{bcolors.PURPLE}Using FedProxAdaptive with base_mu={args.mu} (adaptive per client){bcolors.ENDC}")
    elif args.strategy == "SCAFFOLD":
        strategy = SCAFFOLD(lr_server=1.0)
        print(f"{bcolors.PURPLE}Using SCAFFOLD with server_lr=1.0{bcolors.ENDC}")
    else:
        strategy = FedAvg()
        print(f"{bcolors.PURPLE}Using FedAvg aggregation strategy{bcolors.ENDC}")

    partition_type, client_count = parse_partition_type(args.partition_type)
    n_clients = min(args.n_clients, client_count)
    
    logger, log_filename = setup_logger(n_clients, partition_type, strategy=args.strategy)
    excel_filename = log_filename.replace('.log', '.xlsx')
    results_df = pd.DataFrame(columns=['Round', 'Loss', 'Accuracy', 'F1_Score', 'Precision', 'Recall'])
    
    # Get metadata and class names
    with open(os.path.join("data", "partitions", "meta.txt"), "r") as f:
        num_classes = int(f.read().strip())
    
    class_names = None
    class_names_file = os.path.join("data", "partitions", f"{client_count}_client", partition_type, "label_classes.npy")
    if os.path.exists(class_names_file):
        class_names = np.load(class_names_file, allow_pickle=True)
        class_names = [str(name) for name in class_names]
    
    if class_names is None:
        class_names = [f"Class_{i}" for i in range(num_classes)]
    
    sample_X = np.load(os.path.join("data", "X_test.npy"), mmap_mode='r')
    input_dim = sample_X.shape[1]
    del sample_X
    
    # Setup paths for all clients
    paths_list = []
    for i in range(n_clients):
        paths = setup_paths(str(i), partition_type, client_count)
        paths_list.append(paths)
    
    # Initialize client status for FedAGRU
    client_status = {}
    for i in range(n_clients):
        client_status[i] = {"importance": 1.0, "patience": 0, "active": True}
    
    latest_weights = None
    
    # Main federated learning loop
    for round_num in range(1, args.rounds + 1):
        logger.info(f"Round {round_num}/{args.rounds}")
        print(f"\n{bcolors.HEADER}Round {round_num}/{args.rounds}{bcolors.ENDC}")
        
        # Determine active clients
        active_clients = [i for i in range(n_clients) if client_status[i]["active"]]
        
        if args.strategy == "FedAGRU":
            print(f"{bcolors.OKCYAN}Active clients: {active_clients}{bcolors.ENDC}")
        
        if not active_clients:
            print(f"{bcolors.FAIL}No active clients remaining. Stopping training.{bcolors.ENDC}")
            break
        
        weights_files = []
        sample_sizes = []
        client_losses = []
        participating_clients = []
        
        # Train each active client
        for i in active_clients:
            tf.keras.backend.clear_session()
            gc.collect()
            time.sleep(1)
            
            client_id = str(i)
            print(f"\n{bcolors.BOLD}Processing Client {client_id}{bcolors.ENDC}")
            
            is_gru = args.model_type.lower() in ["gru"]
            model = create_model(input_dim, num_classes, args.model_type, args.batch_size, strategy, i)
            
            if latest_weights is not None:
                model.set_weights(latest_weights)
            
            paths = paths_list[i]
            train_dataset = create_client_dataset(
                paths['train_X'], paths['train_y'], 
                input_dim, num_classes, args.batch_size, is_gru
            )
            
            # Get dataset size
            X_train_mmap = np.load(paths['train_X'], mmap_mode='r')
            X_train_size = X_train_mmap.shape[0]
            sample_sizes.append(X_train_size)
            del X_train_mmap
            
            # Train client model
            print(f"Training model for client {client_id} for {args.epochs} epochs")
            history = model.fit(
                train_dataset,
                epochs=args.epochs,
                callbacks=[
                    tf.keras.callbacks.EarlyStopping(
                        monitor='loss', patience=3, restore_best_weights=True, verbose=1
                    )
                ],
                verbose=1
            )
            
            # Get training loss (debugging)
            if 'loss' in history.history:
                final_loss = float(history.history["loss"][-1])
                initial_loss = float(history.history["loss"][0]) if len(history.history["loss"]) > 0 else "N/A"
                epochs_trained = len(history.history["loss"])
                print(f"Client {client_id} DEBUG: Initial loss={initial_loss}, Final loss={final_loss}, Epochs trained={epochs_trained}")
            else:
                final_loss = 0.0
            
            # Get training loss
            loss = final_loss
            client_losses.append(loss)
            
            # Local evaluation (FedAGRU only)
            local_accuracy = None
            local_loss = None
            if args.strategy == "FedAGRU":
                print(f"Evaluating client {client_id} on test set")
                test_dataset = create_client_dataset(
                    paths['test_X'], paths['test_y'],
                    input_dim, num_classes, args.batch_size, is_gru
                )
                test_results = model.evaluate(test_dataset, verbose=1)
                local_loss = test_results[0]
                local_accuracy = test_results[1]
                del test_dataset
    
            # Save weights
            weights_file = os.path.join(args.weights_cache_dir, f"client_{client_id}_round_{round_num}_weights.pkl")
            
            if args.strategy == "FedAGRU":
                weights_data = {
                    'weights': model.get_weights(),
                    'local_accuracy': local_accuracy,
                    'local_loss': local_loss
                }
                with open(weights_file, 'wb') as f:
                    pickle.dump(weights_data, f)
            elif args.strategy == "SCAFFOLD":
                scaffold_data = model.get_scaffold_update()
                with open(weights_file, 'wb') as f:
                    pickle.dump(scaffold_data, f)
            else:
                with open(weights_file, 'wb') as f:
                    pickle.dump(model.get_weights(), f)
            
            weights_files.append(weights_file)
            participating_clients.append(i)
            
            logger.info(f"Client {client_id}: Training Loss {loss:.4f}, Local Test Loss {local_loss if local_loss else 'N/A'}, Local Accuracy {local_accuracy if local_accuracy else 'N/A'}")
            print(f"{bcolors.WARNING}Client {client_id}: Training completed - Loss {loss:.4f}{bcolors.ENDC}")
            
            del model, train_dataset, history
            aggressive_memory_cleanup()
        
        if not participating_clients:
            print(f"{bcolors.FAIL}No clients participated in this round. Stopping training.{bcolors.ENDC}")
            break
        
        # Load weights from disk for aggregation
        weights_list = []
        local_accuracies_list = []
        local_losses_list = []
        scaffold_updates = []

        for weights_file in weights_files:
            with open(weights_file, 'rb') as f:
                saved_data = pickle.load(f)
                
                if isinstance(saved_data, dict):
                    if 'control_update' in saved_data:  # SCAFFOLD
                        scaffold_updates.append(saved_data)
                        weights_list.append(saved_data['weights'])
                    else:  # FedAGRU
                        weights_list.append(saved_data['weights'])
                        if 'local_accuracy' in saved_data and saved_data['local_accuracy'] is not None:
                            local_accuracies_list.append(saved_data['local_accuracy'])
                        if 'local_loss' in saved_data and saved_data['local_loss'] is not None:
                            local_losses_list.append(saved_data['local_loss'])
                else:
                    weights_list.append(saved_data)

        # Prepare data for aggregation
        local_accuracies_for_aggregation = local_accuracies_list if local_accuracies_list and len(local_accuracies_list) == len(weights_list) else None
        local_losses_for_aggregation = local_losses_list if local_losses_list and len(local_losses_list) == len(weights_list) else None

        print(f"\n{bcolors.OKBLUE}Aggregating weights with {args.strategy}{bcolors.ENDC}")
        
        # Aggregate weights
        if args.strategy == "FedAGRU":
            round_loss = sum(client_losses) / len(client_losses) if client_losses else None
            
            aggregated_weights, importance_vector = strategy.aggregate(
                weights_list, 
                sample_sizes, 
                local_accuracies=local_accuracies_for_aggregation, 
                local_losses=local_losses_for_aggregation,
                round_loss=round_loss
            )
            
            if args.disable_client_selection:
                print(f"{bcolors.CYAN}Client selection DISABLED - using all clients with importance weighting{bcolors.ENDC}")
                for idx, client_id in enumerate(participating_clients):
                    client_status[client_id]["importance"] = importance_vector[idx]
                    client_status[client_id]["patience"] = 0
                    client_status[client_id]["active"] = True
                    logger.info(f"Round {round_num} - Client {client_id}: Importance = {importance_vector[idx]:.4f}")
            else:
                threshold_v = 1.0 / args.n_clients
                
                for idx, client_id in enumerate(participating_clients):
                    new_importance = importance_vector[idx]
                    client_status[client_id]["importance"] = new_importance
                    logger.info(f"Round {round_num} - Client {client_id}: Importance = {new_importance:.4f}")
                    
                    if new_importance >= threshold_v:
                        client_status[client_id]["patience"] = 0
                    else:
                        client_status[client_id]["patience"] += 1
                        if client_status[client_id]["patience"] > args.patience_T:
                            client_status[client_id]["active"] = False
                            print(f"{bcolors.FAIL}Client {client_id} disabled (patience exceeded){bcolors.ENDC}")
        elif args.strategy == "FedAvg2":
            aggregated_weights = strategy.aggregate(weights_list, sample_sizes, client_losses)
        elif args.strategy == "FedProxAdaptive":
            aggregated_weights = strategy.aggregate(weights_list, sample_sizes, client_losses, participating_clients)
        elif args.strategy == "SCAFFOLD":
            aggregated_weights = strategy.aggregate(weights_list, sample_sizes, scaffold_updates)
        else:
            aggregated_weights = strategy.aggregate(weights_list, sample_sizes)
        
        latest_weights = aggregated_weights
        
        # Store weights_list before deletion for FedAGRU update
        weights_list_for_update = weights_list.copy() if args.strategy == "FedAGRU" else None
        del weights_list
        gc.collect()
        
        # Evaluate aggregated model
        eval_model = create_model(input_dim, num_classes, args.model_type, args.batch_size, strategy)
        eval_model.set_weights(aggregated_weights)
        
        test_dataset = create_client_dataset(
            paths_list[0]['test_X'], paths_list[0]['test_y'],
            input_dim, num_classes, args.batch_size, 
            args.model_type.lower() in ["gru"]
        )
        
        print(f"Evaluating aggregated model")
        result = evaluate_model_with_metrics(
            eval_model, test_dataset, num_classes, class_names, 
            round_num, args.strategy, partition_type
        )
        test_loss, accuracy, f1, precision, recall = result[:5]
        per_class_metrics = result[5] if len(result) > 5 else None
        confusion_mat = result[6] if len(result) > 6 else None
        
        # Store results
        new_row = pd.DataFrame({
            'Round': [round_num],
            'Loss': [test_loss],
            'Accuracy': [accuracy],
            'F1_Score': [f1],
            'Precision': [precision],
            'Recall': [recall]
        })
        results_df = pd.concat([results_df, new_row], ignore_index=True)
        results_df.to_excel(excel_filename, index=False)
        
        # Create enhanced Excel file with per-class metrics
        if per_class_metrics is not None:
            create_enhanced_excel_report(excel_filename, results_df, per_class_metrics, 
                                       class_names, round_num, confusion_mat)
        
        # Update FedAGRU learnable parameters with global model performance
        if args.strategy == "FedAGRU" and hasattr(strategy, 'update_from_global_evaluation'):
            strategy.update_from_global_evaluation(weights_list_for_update, importance_vector, test_loss)
        
        round_avg_loss = sum(client_losses) / len(client_losses)
        logger.info(f"Round {round_num} completed - Avg Training Loss: {round_avg_loss:.4f}, Global Test Loss: {test_loss:.4f}, Accuracy: {accuracy*100:.2f}%, F1: {f1:.4f}, Precision: {precision:.4f}, Recall: {recall:.4f}")
        print(f"{bcolors.OKGREEN}Round {round_num} completed - Global Test Loss: {test_loss:.4f}, Accuracy: {accuracy*100:.2f}%, F1: {f1:.4f}, Precision: {precision:.4f}, Recall: {recall:.4f}{bcolors.ENDC}")
        
        # Clean up weights files
        for weights_file in weights_files:
            if round_num != args.rounds:
                os.remove(weights_file)
        
        del eval_model, test_dataset, aggregated_weights
        aggressive_memory_cleanup()
        time.sleep(1)
    
    print(f"{bcolors.OKCYAN}Results saved to {excel_filename}{bcolors.ENDC}")

if __name__ == "__main__":
    main()